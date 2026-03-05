"""All blueprints except auth — dashboard, cases, invoices, clients, team, admin, notifications"""
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, abort, jsonify, send_file)
from flask_login import login_required, current_user
from datetime import datetime, date
from functools import wraps
from app import db
from app.models import (User, Case, Client, TeamMember, Invoice,
                        CaseHistory, Notification, Permission,
                        ROLES, MODULES, ACTIONS, ROLE_DEFAULTS)

# ── Permission decorator ──────────────────────────────────────────────────────
def require(module, action):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.can(module, action):
                abort(403)
            return f(*args, **kwargs)
        return wrapped
    return decorator

# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
dashboard_bp = Blueprint("dashboard", __name__)

@dashboard_bp.route("/")
@login_required
def index():
    today = date.today()
    stats = {
        "total":       Case.query.count(),
        "under_prep":  Case.query.filter_by(status="Under Preparation").count(),
        "prepared":    Case.query.filter_by(status="Prepared").count(),
        "completed":   Case.query.filter(Case.status.in_(["Completed","Case Closed"])).count(),
        "inv_total":   Invoice.query.count(),
        "inv_pending": Invoice.query.filter_by(status="Pending").count(),
        "rev_collected": db.session.query(db.func.sum(Invoice.amount)).filter_by(status="Paid").scalar() or 0,
        "rev_pending":   db.session.query(db.func.sum(Invoice.amount)).filter_by(status="Pending").scalar() or 0,
    }
    by_status   = db.session.query(Case.status, db.func.count()).group_by(Case.status).all()
    by_assignee = (db.session.query(TeamMember.name, db.func.count(Case.id))
                   .join(Case, TeamMember.id == Case.assignee_id)
                   .group_by(TeamMember.name).all())
    by_client   = (db.session.query(Client.name, db.func.count(Case.id))
                   .join(Case, Client.id == Case.client_id)
                   .group_by(Client.name).order_by(db.func.count(Case.id).desc())
                   .limit(8).all())
    due_soon    = (Case.query
                   .filter(Case.compliance_date.isnot(None),
                           Case.compliance_date >= today,
                           Case.status.notin_(["Completed","Case Closed"]))
                   .order_by(Case.compliance_date).limit(10).all())
    overdue     = (Case.query
                   .filter(Case.compliance_date.isnot(None),
                           Case.compliance_date < today,
                           Case.status.notin_(["Completed","Case Closed"]))
                   .order_by(Case.compliance_date).all())
    return render_template("dashboard.html", stats=stats, by_status=by_status,
                           by_assignee=by_assignee, by_client=by_client,
                           due_soon=due_soon, overdue=overdue, today=today)

# ─────────────────────────────────────────────────────────────────────────────
# CASES
# ─────────────────────────────────────────────────────────────────────────────
cases_bp = Blueprint("cases", __name__, url_prefix="/cases")

def _case_context():
    return {
        "clients":  Client.query.order_by(Client.name).all(),
        "team":     TeamMember.query.filter_by(is_active=True).order_by(TeamMember.name).all(),
        "statuses": ["Under Preparation","Prepared","Completed","Case Closed","Under Process","Replied"],
        "priorities": ["High","Normal","Low"],
    }

@cases_bp.route("/")
@login_required
@require("cases","view")
def index():
    q      = request.args.get("q","")
    status = request.args.get("status","All")
    asgn   = request.args.get("assignee","All")
    client = request.args.get("client","All")
    query  = Case.query
    if q:
        query = query.join(Client, isouter=True).join(TeamMember, isouter=True).filter(
            db.or_(Case.case_details.ilike(f"%{q}%"),
                   Case.file_no.ilike(f"%{q}%"),
                   Client.name.ilike(f"%{q}%"),
                   Case.progress.ilike(f"%{q}%")))
    if status != "All": query = query.filter_by(status=status)
    if asgn   != "All":
        tm = TeamMember.query.filter_by(name=asgn).first()
        if tm: query = query.filter_by(assignee_id=tm.id)
    if client != "All":
        cl = Client.query.filter_by(name=client).first()
        if cl: query = query.filter_by(client_id=cl.id)
    cases = query.order_by(Case.updated_at.desc()).all()
    ctx   = _case_context()
    return render_template("cases/index.html", cases=cases, q=q,
                           status=status, asgn=asgn, client=client, **ctx)

@cases_bp.route("/add", methods=["GET","POST"])
@login_required
@require("cases","add")
def add():
    ctx = _case_context()
    if request.method == "POST":
        c = Case(
            file_no         = request.form.get("file_no","").strip(),
            client_id       = request.form.get("client_id") or None,
            case_details    = request.form.get("case_details","").strip(),
            tax_period      = request.form.get("tax_period","").strip(),
            assignee_id     = request.form.get("assignee_id") or None,
            status          = request.form.get("status","Under Preparation"),
            priority        = request.form.get("priority","Normal"),
            compliance_date = _parse_date(request.form.get("compliance_date")),
            completion_date = _parse_date(request.form.get("completion_date")),
            progress        = request.form.get("progress","").strip(),
            notes           = request.form.get("notes","").strip(),
        )
        db.session.add(c)
        db.session.commit()
        flash("Case added successfully.", "success")
        return redirect(url_for("cases.index"))
    return render_template("cases/form.html", case=None, action="Add", **ctx)

@cases_bp.route("/<int:case_id>")
@login_required
@require("cases","view")
def view(case_id):
    case = Case.query.get_or_404(case_id)
    return render_template("cases/view.html", case=case)

@cases_bp.route("/<int:case_id>/edit", methods=["GET","POST"])
@login_required
@require("cases","edit")
def edit(case_id):
    case = Case.query.get_or_404(case_id)
    # Associates can only edit their own cases
    if current_user.role == "associate":
        tm = TeamMember.query.filter_by(user_id=current_user.id).first()
        if not tm or case.assignee_id != tm.id:
            abort(403)
    ctx = _case_context()
    if request.method == "POST":
        _track_changes(case, current_user.full_name or current_user.username)
        case.file_no         = request.form.get("file_no","").strip()
        case.client_id       = request.form.get("client_id") or None
        case.case_details    = request.form.get("case_details","").strip()
        case.tax_period      = request.form.get("tax_period","").strip()
        if current_user.can("cases","edit_assignee"):
            case.assignee_id = request.form.get("assignee_id") or None
        if current_user.can("cases","edit_status"):
            case.status      = request.form.get("status", case.status)
        case.priority        = request.form.get("priority", case.priority)
        if current_user.can("cases","edit_compliance_date"):
            case.compliance_date = _parse_date(request.form.get("compliance_date"))
        if current_user.can("cases","edit_completion_date"):
            case.completion_date = _parse_date(request.form.get("completion_date"))
        if current_user.can("cases","edit_progress"):
            case.progress    = request.form.get("progress","").strip()
        case.notes           = request.form.get("notes","").strip()
        case.updated_at      = datetime.utcnow()
        db.session.commit()
        flash("Case updated.", "success")
        return redirect(url_for("cases.view", case_id=case.id))
    return render_template("cases/form.html", case=case, action="Edit", **ctx)

@cases_bp.route("/<int:case_id>/delete", methods=["POST"])
@login_required
@require("cases","delete")
def delete(case_id):
    case = Case.query.get_or_404(case_id)
    db.session.delete(case)
    db.session.commit()
    flash("Case deleted.", "warning")
    return redirect(url_for("cases.index"))

def _track_changes(case, user):
    fields = ["file_no","case_details","tax_period","status","priority",
              "compliance_date","completion_date","progress"]
    for f in fields:
        old = str(getattr(case, f) or "")
        # Will be compared after form submission — simplified tracking
        h = CaseHistory(case_id=case.id, field=f, old_value=old,
                        new_value="(pending)", changed_by=user)
        db.session.add(h)

def _parse_date(s):
    if not s: return None
    try:    return datetime.strptime(s, "%Y-%m-%d").date()
    except: return None

# ─────────────────────────────────────────────────────────────────────────────
# INVOICES
# ─────────────────────────────────────────────────────────────────────────────
invoices_bp = Blueprint("invoices", __name__, url_prefix="/invoices")

@invoices_bp.route("/")
@login_required
@require("invoices","view")
def index():
    status = request.args.get("status","All")
    client = request.args.get("client","All")
    query  = Invoice.query
    if status != "All": query = query.filter_by(status=status)
    if client != "All":
        cl = Client.query.filter_by(name=client).first()
        if cl: query = query.filter_by(client_id=cl.id)
    invoices = query.order_by(Invoice.created_at.desc()).all()
    clients  = Client.query.order_by(Client.name).all()
    total    = sum(i.amount or 0 for i in invoices)
    paid     = sum(i.amount or 0 for i in invoices if i.status=="Paid")
    pending  = sum(i.amount or 0 for i in invoices if i.status=="Pending")
    return render_template("invoices/index.html", invoices=invoices,
                           clients=clients, status=status, client=client,
                           total=total, paid=paid, pending=pending)

@invoices_bp.route("/add", methods=["GET","POST"])
@login_required
@require("invoices","add")
def add():
    if request.method == "POST":
        inv = Invoice(
            invoice_no  = request.form.get("invoice_no","").strip(),
            case_id     = request.form.get("case_id") or None,
            client_id   = request.form.get("client_id") or None,
            description = request.form.get("description","").strip(),
            amount      = float(request.form.get("amount",0) or 0),
            status      = request.form.get("status","Pending"),
            issue_date  = _parse_date(request.form.get("issue_date")),
            due_date    = _parse_date(request.form.get("due_date")),
            paid_date   = _parse_date(request.form.get("paid_date")),
            notes       = request.form.get("notes","").strip(),
        )
        db.session.add(inv)
        db.session.commit()
        flash("Invoice added.", "success")
        return redirect(url_for("invoices.index"))
    clients = Client.query.order_by(Client.name).all()
    cases   = Case.query.order_by(Case.case_details).all()
    return render_template("invoices/form.html", invoice=None, action="Add",
                           clients=clients, cases=cases,
                           statuses=["Pending","Paid","Partial","Cancelled"])

@invoices_bp.route("/<int:inv_id>/edit", methods=["GET","POST"])
@login_required
@require("invoices","edit")
def edit(inv_id):
    inv = Invoice.query.get_or_404(inv_id)
    if request.method == "POST":
        inv.invoice_no  = request.form.get("invoice_no","").strip()
        inv.case_id     = request.form.get("case_id") or None
        inv.client_id   = request.form.get("client_id") or None
        inv.description = request.form.get("description","").strip()
        inv.amount      = float(request.form.get("amount",0) or 0)
        if current_user.can("invoices","edit_invoice_status"):
            inv.status  = request.form.get("status", inv.status)
        inv.issue_date  = _parse_date(request.form.get("issue_date"))
        inv.due_date    = _parse_date(request.form.get("due_date"))
        inv.paid_date   = _parse_date(request.form.get("paid_date"))
        inv.notes       = request.form.get("notes","").strip()
        db.session.commit()
        flash("Invoice updated.", "success")
        return redirect(url_for("invoices.index"))
    clients = Client.query.order_by(Client.name).all()
    cases   = Case.query.order_by(Case.case_details).all()
    return render_template("invoices/form.html", invoice=inv, action="Edit",
                           clients=clients, cases=cases,
                           statuses=["Pending","Paid","Partial","Cancelled"])

@invoices_bp.route("/<int:inv_id>/delete", methods=["POST"])
@login_required
@require("invoices","delete")
def delete(inv_id):
    inv = Invoice.query.get_or_404(inv_id)
    db.session.delete(inv)
    db.session.commit()
    flash("Invoice deleted.", "warning")
    return redirect(url_for("invoices.index"))

# ─────────────────────────────────────────────────────────────────────────────
# CLIENTS
# ─────────────────────────────────────────────────────────────────────────────
clients_bp = Blueprint("clients", __name__, url_prefix="/clients")

@clients_bp.route("/")
@login_required
@require("clients","view")
def index():
    q = request.args.get("q","")
    query = Client.query
    if q: query = query.filter(Client.name.ilike(f"%{q}%"))
    clients = query.order_by(Client.name).all()
    return render_template("clients/index.html", clients=clients, q=q)

@clients_bp.route("/add", methods=["GET","POST"])
@login_required
@require("clients","add")
def add():
    if request.method == "POST":
        c = Client(name=request.form.get("name","").strip(),
                   ntn=request.form.get("ntn","").strip(),
                   contact=request.form.get("contact","").strip(),
                   email=request.form.get("email","").strip(),
                   address=request.form.get("address","").strip(),
                   notes=request.form.get("notes","").strip())
        db.session.add(c)
        db.session.commit()
        flash("Client added.", "success")
        return redirect(url_for("clients.index"))
    return render_template("clients/form.html", client=None, action="Add")

@clients_bp.route("/<int:cid>/edit", methods=["GET","POST"])
@login_required
@require("clients","edit")
def edit(cid):
    c = Client.query.get_or_404(cid)
    if request.method == "POST":
        c.name=request.form.get("name","").strip(); c.ntn=request.form.get("ntn","").strip()
        c.contact=request.form.get("contact","").strip(); c.email=request.form.get("email","").strip()
        c.address=request.form.get("address","").strip(); c.notes=request.form.get("notes","").strip()
        db.session.commit(); flash("Client updated.", "success")
        return redirect(url_for("clients.index"))
    return render_template("clients/form.html", client=c, action="Edit")

@clients_bp.route("/<int:cid>/delete", methods=["POST"])
@login_required
@require("clients","delete")
def delete(cid):
    c = Client.query.get_or_404(cid)
    db.session.delete(c); db.session.commit()
    flash("Client deleted.", "warning")
    return redirect(url_for("clients.index"))

# ─────────────────────────────────────────────────────────────────────────────
# TEAM
# ─────────────────────────────────────────────────────────────────────────────
team_bp = Blueprint("team", __name__, url_prefix="/team")

@team_bp.route("/")
@login_required
@require("team","view")
def index():
    members = TeamMember.query.order_by(TeamMember.name).all()
    case_counts = {m.id: len(m.cases) for m in members}
    return render_template("team/index.html", members=members, case_counts=case_counts)

@team_bp.route("/add", methods=["GET","POST"])
@login_required
@require("team","add")
def add():
    if request.method == "POST":
        m = TeamMember(name=request.form.get("name","").strip(),
                       role=request.form.get("role","").strip(),
                       email=request.form.get("email","").strip(),
                       phone=request.form.get("phone","").strip(),
                       is_active=bool(request.form.get("is_active")))
        db.session.add(m); db.session.commit()
        flash("Team member added.", "success")
        return redirect(url_for("team.index"))
    return render_template("team/form.html", member=None, action="Add")

@team_bp.route("/<int:mid>/edit", methods=["GET","POST"])
@login_required
@require("team","edit")
def edit(mid):
    m = TeamMember.query.get_or_404(mid)
    if request.method == "POST":
        m.name=request.form.get("name","").strip(); m.role=request.form.get("role","").strip()
        m.email=request.form.get("email","").strip(); m.phone=request.form.get("phone","").strip()
        m.is_active=bool(request.form.get("is_active"))
        db.session.commit(); flash("Member updated.", "success")
        return redirect(url_for("team.index"))
    return render_template("team/form.html", member=m, action="Edit")

@team_bp.route("/<int:mid>/delete", methods=["POST"])
@login_required
@require("team","delete")
def delete(mid):
    m = TeamMember.query.get_or_404(mid)
    db.session.delete(m); db.session.commit()
    flash("Member removed.", "warning")
    return redirect(url_for("team.index"))

# ─────────────────────────────────────────────────────────────────────────────
# NOTIFICATIONS
# ─────────────────────────────────────────────────────────────────────────────
notif_bp = Blueprint("notifications", __name__, url_prefix="/notifications")

@notif_bp.route("/")
@login_required
def index():
    notifs = Notification.query.filter_by(user_id=current_user.id)\
               .order_by(Notification.created_at.desc()).limit(50).all()
    return render_template("notifications/index.html", notifications=notifs)

@notif_bp.route("/<int:nid>/read")
@login_required
def mark_read(nid):
    n = Notification.query.get_or_404(nid)
    if n.user_id == current_user.id:
        n.is_read = True; db.session.commit()
    return redirect(n.link or url_for("notifications.index"))

@notif_bp.route("/mark-all-read", methods=["POST"])
@login_required
def mark_all_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False)\
        .update({"is_read": True})
    db.session.commit()
    return redirect(url_for("notifications.index"))

@notif_bp.route("/api/unread-count")
@login_required
def unread_count():
    return jsonify({"count": current_user.unread_count()})

# ─────────────────────────────────────────────────────────────────────────────
# ADMIN — User & Permission Management
# ─────────────────────────────────────────────────────────────────────────────
admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

def admin_only(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if current_user.role != "admin":
            abort(403)
        return f(*args, **kwargs)
    return wrapped

@admin_bp.route("/users")
@login_required
@admin_only
def users():
    all_users = User.query.order_by(User.created_at.desc()).all()
    return render_template("admin/users.html", users=all_users, roles=ROLES)

@admin_bp.route("/users/add", methods=["GET","POST"])
@login_required
@admin_only
def add_user():
    error = None
    if request.method == "POST":
        from werkzeug.security import generate_password_hash
        username = request.form.get("username","").strip()
        email    = request.form.get("email","").strip()
        if User.query.filter_by(username=username).first():
            error = "Username already exists."
        elif User.query.filter_by(email=email).first():
            error = "Email already exists."
        else:
            u = User(username=username, email=email,
                     password_hash=generate_password_hash(request.form.get("password","")),
                     full_name=request.form.get("full_name","").strip(),
                     role=request.form.get("role","viewer"),
                     is_active=bool(request.form.get("is_active")))
            db.session.add(u); db.session.commit()
            flash(f"User '{username}' created.", "success")
            return redirect(url_for("admin.users"))
    return render_template("admin/user_form.html", user=None, action="Add",
                           roles=ROLES, error=error)

@admin_bp.route("/users/<int:uid>/edit", methods=["GET","POST"])
@login_required
@admin_only
def edit_user(uid):
    u = User.query.get_or_404(uid)
    error = None
    if request.method == "POST":
        u.full_name = request.form.get("full_name","").strip()
        u.role      = request.form.get("role", u.role)
        u.is_active = bool(request.form.get("is_active"))
        new_pwd     = request.form.get("new_password","")
        if new_pwd:
            from werkzeug.security import generate_password_hash
            if len(new_pwd) < 8:
                error = "Password must be at least 8 characters."
            else:
                u.password_hash = generate_password_hash(new_pwd)
        if not error:
            db.session.commit()
            flash(f"User '{u.username}' updated.", "success")
            return redirect(url_for("admin.permissions", uid=u.id))
    return render_template("admin/user_form.html", user=u, action="Edit",
                           roles=ROLES, error=error)

@admin_bp.route("/users/<int:uid>/delete", methods=["POST"])
@login_required
@admin_only
def delete_user(uid):
    if uid == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin.users"))
    u = User.query.get_or_404(uid)
    db.session.delete(u); db.session.commit()
    flash("User deleted.", "warning")
    return redirect(url_for("admin.users"))

@admin_bp.route("/users/<int:uid>/permissions", methods=["GET","POST"])
@login_required
@admin_only
def permissions(uid):
    u = User.query.get_or_404(uid)
    if request.method == "POST":
        # Delete existing custom permissions
        Permission.query.filter_by(user_id=u.id).delete()
        for module in MODULES:
            for action in ACTIONS:
                key     = f"{module}__{action}"
                allowed = bool(request.form.get(key))
                p = Permission(user_id=u.id, module=module,
                               action=action, allowed=allowed)
                db.session.add(p)
        db.session.commit()
        flash(f"Permissions saved for {u.full_name or u.username}.", "success")
        return redirect(url_for("admin.users"))

    # Build current permission matrix
    existing = {(p.module, p.action): p.allowed for p in u.permissions}
    matrix   = {}
    for m in MODULES:
        matrix[m] = {}
        for a in ACTIONS:
            if (m, a) in existing:
                matrix[m][a] = existing[(m, a)]
            else:
                matrix[m][a] = ROLE_DEFAULTS.get(u.role, {}).get(m, {}).get(a, False)
    return render_template("admin/permissions.html", user=u, matrix=matrix,
                           modules=MODULES, actions=ACTIONS)

@admin_bp.route("/users/<int:uid>/reset-to-role", methods=["POST"])
@login_required
@admin_only
def reset_to_role(uid):
    u = User.query.get_or_404(uid)
    Permission.query.filter_by(user_id=u.id).delete()
    db.session.commit()
    flash(f"Permissions reset to role defaults for {u.full_name or u.username}.", "info")
    return redirect(url_for("admin.permissions", uid=u.id))

# ─────────────────────────────────────────────────────────────────────────────
# IMPORT
# ─────────────────────────────────────────────────────────────────────────────
import_bp = Blueprint("import_data", __name__, url_prefix="/import")

@import_bp.route("/", methods=["GET", "POST"])
@login_required
@admin_only
def index():
    if request.method == "POST":
        f = request.files.get("excel_file")
        if not f or not f.filename.endswith((".xlsx", ".xls")):
            flash("Please upload a valid Excel file (.xlsx or .xls).", "danger")
            return redirect(url_for("import_data.index"))

        try:
            import openpyxl, io
            wb   = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            sheet_name = request.form.get("sheet_name", "").strip()

            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            rows    = list(ws.iter_rows(values_only=True))
            if not rows:
                flash("The selected sheet is empty.", "danger")
                return redirect(url_for("import_data.index"))

            # Detect headers from first row
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]

            def col(fragments):
                for i, h in enumerate(headers):
                    if any(frag in h for frag in fragments):
                        return i
                return None

            ci_client   = col(["client"])
            ci_details  = col(["case detail", "case"])
            ci_period   = col(["tax period", "period"])
            ci_assignee = col(["assignee"])
            ci_fileno   = col(["file no", "file"])
            ci_status   = col(["status"])
            ci_progress = col(["progress"])
            ci_invoice  = col(["invoice"])
            ci_complete = col(["completion", "expected"])

            if ci_client is None or ci_details is None:
                flash(f"Could not find 'Client' or 'Case Details' columns. "
                      f"Found: {', '.join(h for h in headers if h)}", "danger")
                return redirect(url_for("import_data.index"))

            # Build lookups
            client_lookup = {c.name.lower(): c.id for c in Client.query.all()}
            team_lookup   = {m.name.lower(): m.id for m in TeamMember.query.all()}

            imported = skipped = 0
            for row in rows[1:]:
                def val(ci):
                    if ci is None or ci >= len(row): return ""
                    v = row[ci]
                    return str(v).strip() if v is not None else ""

                client_name  = val(ci_client)
                case_details = val(ci_details)

                if not client_name or not case_details or case_details.lower() == "nan":
                    skipped += 1
                    continue

                # Auto-create client if not exists
                cl_key = client_name.lower()
                if cl_key not in client_lookup:
                    new_cl = Client(name=client_name)
                    db.session.add(new_cl)
                    db.session.flush()
                    client_lookup[cl_key] = new_cl.id

                # Auto-create team member if not exists
                assignee_name = val(ci_assignee)
                a_key = assignee_name.lower()
                if a_key and a_key not in team_lookup:
                    new_tm = TeamMember(name=assignee_name, role="Associate", is_active=True)
                    db.session.add(new_tm)
                    db.session.flush()
                    team_lookup[a_key] = new_tm.id

                status = val(ci_status) if ci_status is not None else ""
                if not status or status.lower() == "nan":
                    status = "Under Preparation"

                # Parse completion date
                comp_date = None
                if ci_complete is not None:
                    raw = val(ci_complete)
                    if raw and raw.lower() != "nan":
                        try:
                            from datetime import datetime as dt
                            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                                try:
                                    comp_date = dt.strptime(raw, fmt).date(); break
                                except: pass
                        except: pass

                case = Case(
                    file_no         = val(ci_fileno),
                    client_id       = client_lookup.get(cl_key),
                    case_details    = case_details,
                    tax_period      = val(ci_period),
                    assignee_id     = team_lookup.get(a_key) if a_key else None,
                    status          = status,
                    priority        = "Normal",
                    completion_date = comp_date,
                    progress        = val(ci_progress) if ci_progress is not None else "",
                    notes           = val(ci_invoice)  if ci_invoice  is not None else "",
                )
                db.session.add(case)
                imported += 1

            db.session.commit()
            flash(f"✅ Successfully imported {imported} cases. "
                  f"({skipped} empty rows skipped)", "success")
            return redirect(url_for("cases.index"))

        except Exception as e:
            db.session.rollback()
            flash(f"Import failed: {str(e)}", "danger")
            return redirect(url_for("import_data.index"))

    # GET — show form with sheet names if file already chosen
    return render_template("import/index.html")
